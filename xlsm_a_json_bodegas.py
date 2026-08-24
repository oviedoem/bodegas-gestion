"""
xlsm_a_json_bodegas.py
Lee datos-bodegas.xlsm (generado por modBodegas.BajarTodoBat) y genera
bodegas_gestion.json y bodegas_ir_otras.json con la misma estructura
que los scripts Python directos.

Lógica incluida:
  - diasAntiguedad = (hoy - FECHA_EMISION).days  (formula exacta del panel-admin)
  - _deduplicar_y_acumular: GRT manda; GIB/GME mismo dia que GRT se excluyen;
    acumula hasta cubrir ST_FISICO; conserva el más reciente por codigo
  - regla anti-retroceso: aborta si nueva descarga < 50% de la anterior
"""
import json, datetime, sys
from pathlib import Path
from collections import defaultdict

import openpyxl  # pip install openpyxl

BASE_DIR  = Path(__file__).parent
XLSM_PATH = BASE_DIR / "datos-bodegas.xlsm"
OUT_GESTION = BASE_DIR / "bodegas_gestion.json"
OUT_IR      = BASE_DIR / "bodegas_ir_otras.json"

# Columnas del XLSM (posicion 1-based igual que cabeceras en modBodegas.bas)
COL = {
    "SUCURSAL_TAB": 1, "IDBODEGA": 2, "BODEGA": 3, "TIPO_DOC": 4, "FOLIO": 5,
    "CODIGO_TECNICO": 6, "DESCRIPCION": 7,
    "STOCK_DISPONIBLE": 8, "STOCK_FISICO": 9, "CANTIDAD_DOC": 10,
    "FECHA_EMISION": 11, "OBSERVACION_IMPRESA": 12, "COSTO_PROMEDIO": 13,
    "FECHA_REGISTRO_SISTEMA": 14, "USUARIO": 15, "ESTACION_PC": 16,
    "HIPERFAMILIA": 17, "FAMILIA": 18, "SUBFAMILIA": 19, "MARCA": 20,
}

DOC_NOMBRES = {
    "GRT": "Guía Recepción Traslado", "GIB": "Guía Ingreso Entre Bodegas",
    "GII": "Guía Ingreso Inventario", "GME": "Guía Elect. Despacho Factura",
    "Gdc": "Guía Devolución Cliente", "GRC": "Guía Recepción Compra",
    "GTS": "Guía Traslados Entre Sucursales", "GST": "Solicitud de Traslado",
    "GEI": "Guía Egreso Inventario / Merma-Gestión", "GDV": "Guía Despacho Venta",
}

# Definicion de grupos y su hoja XLSM
GRUPOS = [
    {
        "hoja": "BOD_EM", "idSucursal": "04", "nombre": "El Manzano",
        "json": "gestion",
        "bodegas": [
            {"id": 28, "simbolo": "GEM", "nombre": "Gestion El Manzano",   "idsucursal": "04"},
            {"id": 29, "simbolo": "MEM", "nombre": "Mermas El Manzano",     "idsucursal": "04"},
            {"id": 55, "simbolo": "RCE", "nombre": "Recepcion El Manzano",  "idsucursal": "04"},
            {"id": 72, "simbolo": "IEM", "nombre": "Ingreso El Manzano",    "idsucursal": "04"},
            {"id": 46, "simbolo": "TEM", "nombre": "Transito El Manzano",   "idsucursal": "04"},
            {"id": 83, "simbolo": "EEM", "nombre": "Exhibicion El Manzano", "idsucursal": "04"},
        ],
    },
    {
        "hoja": "BOD_SV", "idSucursal": "05", "nombre": "San Vicente",
        "json": "gestion",
        "bodegas": [
            {"id": 41, "simbolo": "GSV", "nombre": "Gestion San Vicente",      "idsucursal": "05"},
            {"id": 42, "simbolo": "MSV", "nombre": "Mermas San Vicente",        "idsucursal": "05"},
            {"id": 56, "simbolo": "RSV", "nombre": "Recepcion San Vicente",     "idsucursal": "05"},
            {"id": 70, "simbolo": "ISV", "nombre": "Ingreso San Vicente",       "idsucursal": "05"},
            {"id": 45, "simbolo": "TSV", "nombre": "Transito San Vicente",      "idsucursal": "05"},
            {"id": 44, "simbolo": "CSV", "nombre": "Calzada San Vicente",       "idsucursal": "05"},
            {"id": 88, "simbolo": "DSV", "nombre": "Distribucion San Vicente",  "idsucursal": "14"},
            {"id": 95, "simbolo": "ESV", "nombre": "Exhibicion San Vicente",    "idsucursal": "05"},
            {"id": 43, "simbolo": "CSV", "nombre": "Consumo San Vicente",       "idsucursal": "05"},
        ],
    },
    {
        "hoja": "BOD_LC", "idSucursal": "06", "nombre": "Las Cabras",
        "json": "gestion",
        "bodegas": [
            {"id": 37, "simbolo": "GLC", "nombre": "Gestion Las Cabras",  "idsucursal": "06"},
            {"id": 38, "simbolo": "MLC", "nombre": "Mermas Las Cabras",    "idsucursal": "06"},
            {"id": 57, "simbolo": "RLC", "nombre": "Recepcion Las Cabras", "idsucursal": "06"},
            {"id": 71, "simbolo": "ILC", "nombre": "Ingreso Las Cabras",   "idsucursal": "06"},
            {"id": 16, "simbolo": "TLC", "nombre": "Transito Las Cabras",  "idsucursal": "06"},
            {"id": 35, "simbolo": "CLC", "nombre": "Calzada Las Cabras",   "idsucursal": "06"},
            {"id": 91, "simbolo": "GFL", "nombre": "Garantia Las Cabras",  "idsucursal": "06"},
            {"id": 96, "simbolo": "ELC", "nombre": "Exhibicion Las Cabras","idsucursal": "06"},
            {"id": 97, "simbolo": "VLC", "nombre": "Volumen Las Cabras",   "idsucursal": "06"},
        ],
    },
    {
        "hoja": "BOD_LT", "idSucursal": "11", "nombre": "Litueche",
        "json": "gestion",
        "bodegas": [
            {"id": 63, "simbolo": "GLE", "nombre": "Gestion Litueche",    "idsucursal": "11"},
            {"id": 76, "simbolo": "MLE", "nombre": "Mermas Litueche",      "idsucursal": "11"},
            {"id": 74, "simbolo": "ILE", "nombre": "Ingreso Litueche",     "idsucursal": "11"},
            {"id": 59, "simbolo": "TLE", "nombre": "Transito Litueche",    "idsucursal": "11"},
            {"id": 78, "simbolo": "CLT", "nombre": "Calzada Litueche",     "idsucursal": "11"},
            {"id": 79, "simbolo": "DLT", "nombre": "Distribucion Litueche","idsucursal": "09"},
            {"id": 64, "simbolo": "ELE", "nombre": "Exhibicion Litueche",  "idsucursal": "11"},
        ],
    },
    {
        "hoja": "BOD_IR", "idSucursal": "02", "nombre": "Isabel Riquelme",
        "json": "ir",
        "bodegas": [
            {"id":  5, "simbolo": "CAL", "nombre": "Calzada",                    "idsucursal": "02"},
            {"id":  6, "simbolo": "SER", "nombre": "Servicio Tecnico",            "idsucursal": "02"},
            {"id": 25, "simbolo": "WEB", "nombre": "Retiro Web Santiago",         "idsucursal": "02"},
            {"id": 30, "simbolo": "GO",  "nombre": "Gestion Isabel Riquelme",     "idsucursal": "02"},
            {"id": 53, "simbolo": "GAR", "nombre": "Garantia Santiago",           "idsucursal": "02"},
            {"id": 69, "simbolo": "IIR", "nombre": "Ingreso Isabel Riquelme",     "idsucursal": "02"},
            {"id": 77, "simbolo": "BMC", "nombre": "Marticorena Stgo",            "idsucursal": "02"},
            {"id": 92, "simbolo": "RST", "nombre": "Recepcion Santiago",          "idsucursal": "02"},
            {"id": 99, "simbolo": "HEL", "nombre": "Herramientas Electricas",     "idsucursal": "02"},
            {"id": 85, "simbolo": "EIR", "nombre": "Exhibicion Isabel Riquelme",  "idsucursal": "02"},
        ],
    },
    {
        "hoja": "BOD_CD", "idSucursal": "COMPARTIDAS", "nombre": "Compartidas/CD",
        "json": "gestion",
        "bodegas": [
            {"id": 23, "simbolo": "CD",  "nombre": "Centro de Distribucion",         "idsucursal": "08"},
            {"id":  7, "simbolo": "XCD", "nombre": "CrossDock Centro Distribucion",  "idsucursal": "08"},
            {"id": 27, "simbolo": "GCD", "nombre": "Gestion CD",                     "idsucursal": "08"},
            {"id": 73, "simbolo": "ICD", "nombre": "Ingreso Centro Distribucion",    "idsucursal": "08"},
            {"id": 26, "simbolo": "MCD", "nombre": "Mermas CD",                      "idsucursal": "08"},
            {"id": 54, "simbolo": "RCD", "nombre": "Recepcion Centro Distribucion",  "idsucursal": "08"},
            {"id": 67, "simbolo": "TCD", "nombre": "Transito Centro Distribucion",   "idsucursal": "08"},
            {"id": 98, "simbolo": "BDP", "nombre": "Despacho Proveedor",             "idsucursal": "09"},
            {"id": 84, "simbolo": "REM", "nombre": "Remate",                         "idsucursal": "01"},
            {"id": 36, "simbolo": "MKT", "nombre": "Marketing",                      "idsucursal": "01"},
        ],
    },
]


def fecha_str(v):
    if v is None: return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    return str(v)


def fecha_iso(v):
    if v is None: return ""
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date): return v.isoformat()
    return ""


def _deduplicar_y_acumular(registros):
    """
    Misma logica que generar_bodegas_gestion.py (y panel-admin descargar_bod.py).
    """
    grupos = defaultdict(list)
    for r in registros:
        grupos[r["codigoTecnico"]].append(r)

    resultado = []
    for cod, docs in grupos.items():
        total_fisico = docs[0].get("fisico", 0) if docs else 0
        if total_fisico == 0:
            continue

        grt_fechas = {d["_fechaRaw"] for d in docs if d["tipoDoc"] == "GRT"}

        dedup = []
        for doc in docs:
            tipo = doc["tipoDoc"]
            fecha = doc["_fechaRaw"]
            if tipo == "GRT":
                dedup.append(doc)
            elif tipo in ("GME", "GIB"):
                if fecha in grt_fechas:
                    continue
                dedup.append(doc)
            else:
                dedup.append(doc)

        if total_fisico > 0:
            acum = 0
            for doc in dedup:
                if acum >= total_fisico:
                    break
                acum += doc.get("cantidad", 0)
                resultado.append(doc)
        else:
            if dedup:
                resultado.append(dedup[0])

    visto = {}
    for doc in resultado:
        cod = doc["codigoTecnico"]
        dias = doc.get("diasAntiguedad") if doc.get("diasAntiguedad") is not None else 999999
        prev = visto[cod].get("diasAntiguedad") if cod in visto and visto[cod].get("diasAntiguedad") is not None else 999999
        if cod not in visto or dias < prev:
            visto[cod] = doc

    for r in visto.values():
        r.pop("_fechaRaw", None)

    return list(visto.values())


def leer_hoja(ws, hoja_nombre):
    """Lee una hoja del XLSM y devuelve lista de registros crudos."""
    hoy = datetime.date.today()
    crudos = []
    primera = True
    for row in ws.iter_rows(values_only=True):
        if primera:
            primera = False
            continue  # saltar cabecera
        if not row[COL["CODIGO_TECNICO"] - 1]:
            continue
        bodega    = str(row[COL["BODEGA"] - 1] or "").strip()
        tipo_doc  = str(row[COL["TIPO_DOC"] - 1] or "").strip()
        folio     = str(row[COL["FOLIO"] - 1] or "").strip()
        cod_tec   = str(row[COL["CODIGO_TECNICO"] - 1] or "").strip()
        desc      = str(row[COL["DESCRIPCION"] - 1] or "").strip()
        disp      = float(row[COL["STOCK_DISPONIBLE"] - 1] or 0)
        fisico    = float(row[COL["STOCK_FISICO"] - 1] or 0)
        cantidad  = float(row[COL["CANTIDAD_DOC"] - 1] or 0)
        fecha_em  = row[COL["FECHA_EMISION"] - 1]
        obs       = str(row[COL["OBSERVACION_IMPRESA"] - 1] or "").strip().replace("_x000D_", "").strip()
        costo     = round(float(row[COL["COSTO_PROMEDIO"] - 1] or 0))
        fecha_sis = str(row[COL["FECHA_REGISTRO_SISTEMA"] - 1] or "").strip()
        usuario   = str(row[COL["USUARIO"] - 1] or "").strip()
        estacion  = str(row[COL["ESTACION_PC"] - 1] or "").strip()
        hiper     = str(row[COL["HIPERFAMILIA"] - 1] or "").strip()
        fam       = str(row[COL["FAMILIA"] - 1] or "").strip()
        sub       = str(row[COL["SUBFAMILIA"] - 1] or "").strip()
        marca     = str(row[COL["MARCA"] - 1] or "").strip()
        idbodega  = int(row[COL["IDBODEGA"] - 1] or 0)
        suc_tab   = str(row[COL["SUCURSAL_TAB"] - 1] or "").strip()

        # diasAntiguedad: formula exacta panel-admin = (hoy - fecha_emision).days
        if isinstance(fecha_em, datetime.datetime):
            fecha_date = fecha_em.date()
        elif isinstance(fecha_em, datetime.date):
            fecha_date = fecha_em
        else:
            fecha_date = None

        dias     = (hoy - fecha_date).days if fecha_date else None
        fecha_raw = fecha_date.isoformat() if fecha_date else ""

        crudos.append({
            "bodega": bodega, "bodegaNombre": bodega,
            "tipoDoc": tipo_doc,
            "tipoDocNombre": DOC_NOMBRES.get(tipo_doc, tipo_doc),
            "folio": folio, "codigoTecnico": cod_tec, "descripcion": desc,
            "disp": disp, "fisico": fisico, "cantidad": cantidad, "costo": costo,
            "fechaRegistro": fecha_str(fecha_em),
            "fechaRegistroIso": fecha_iso(fecha_em),
            "diasAntiguedad": dias, "observacion": obs,
            "usuario": usuario, "estacionPc": estacion,
            "fechaRegistroSistema": fecha_sis,
            "hiperfamilia": hiper, "familia": fam, "subfamilia": sub, "marca": marca,
            "_fechaRaw": fecha_raw,
            "_idbodega": idbodega, "_sucursalTab": suc_tab,
        })
    return crudos


def anti_retroceso(out_path, nuevo_total, etiqueta):
    if out_path.exists():
        try:
            ant = json.loads(out_path.read_text(encoding="utf-8"))
            ant_total = ant.get("total", 0)
            if ant_total > 0 and nuevo_total < ant_total * 0.5:
                print(f"[ABORTADO] {etiqueta}: {nuevo_total} registros vs {ant_total} anteriores "
                      f"(caida >50%). Se conserva reporte anterior.")
                return False
        except Exception as e:
            print(f"[AVISO] No se pudo leer JSON anterior para anti-retroceso: {e}")
    return True


def main():
    if not XLSM_PATH.exists():
        print(f"[ERROR] No encontrado: {XLSM_PATH}")
        print("Ejecutar CREAR_BODEGAS_XLSM.vbs + DESCARGAR_BODEGAS.bat primero.")
        sys.exit(1)

    print(f"Abriendo {XLSM_PATH.name} (read-only)...")
    wb = openpyxl.load_workbook(str(XLSM_PATH), read_only=True, data_only=True, keep_vba=False)
    print("OK")

    hoy = datetime.date.today()

    # ── Acumular por destino JSON ──────────────────────────────────────────────
    sucursales_gestion = []
    compartidas_registros = []
    total_gestion = 0

    bodegas_ir = []
    total_ir = 0

    for grupo in GRUPOS:
        hoja_nombre = grupo["hoja"]
        if hoja_nombre not in wb.sheetnames:
            print(f"[AVISO] Hoja {hoja_nombre} no encontrada en XLSM — se omite")
            continue

        ws = wb[hoja_nombre]
        print(f"Leyendo {hoja_nombre}...", end=" ", flush=True)
        crudos = leer_hoja(ws, hoja_nombre)
        print(f"{len(crudos)} filas crudas")

        # Dedup por bodega (idbodega)
        por_bodega = defaultdict(list)
        for r in crudos:
            por_bodega[r["_idbodega"]].append(r)

        if grupo["json"] == "ir":
            # Bodegas Isabel Riquelme otras → bodegas_ir_otras.json
            registros_grupo = []
            for bod in grupo["bodegas"]:
                recs = por_bodega.get(bod["id"], [])
                dedup, _ = _deduplicar_y_acumular(recs), len(recs)
                # _deduplicar_y_acumular returns list
                dedup = _deduplicar_y_acumular(recs)
                # Asignar bodegaNombre correcto
                for r in dedup:
                    r["bodega"] = bod["simbolo"]
                    r["bodegaNombre"] = bod["nombre"]
                registros_grupo.extend(dedup)
                print(f"  {bod['simbolo']:5s} {bod['nombre']:30s} → {len(dedup)} codigos")
            bodegas_ir.extend(registros_grupo)
            total_ir += len(registros_grupo)

        elif grupo["idSucursal"] == "COMPARTIDAS":
            registros_grupo = []
            for bod in grupo["bodegas"]:
                recs = por_bodega.get(bod["id"], [])
                dedup = _deduplicar_y_acumular(recs)
                for r in dedup:
                    r["bodega"] = bod["simbolo"]
                    r["bodegaNombre"] = bod["nombre"]
                registros_grupo.extend(dedup)
                print(f"  {bod['simbolo']:5s} {bod['nombre']:30s} → {len(dedup)} codigos")
            compartidas_registros = registros_grupo
            total_gestion += len(registros_grupo)

        else:
            registros_suc = []
            for bod in grupo["bodegas"]:
                recs = por_bodega.get(bod["id"], [])
                dedup = _deduplicar_y_acumular(recs)
                for r in dedup:
                    r["bodega"] = bod["simbolo"]
                    r["bodegaNombre"] = bod["nombre"]
                registros_suc.extend(dedup)
                print(f"  {bod['simbolo']:5s} {bod['nombre']:30s} → {len(dedup)} codigos")
            registros_suc.sort(
                key=lambda r: r.get("diasAntiguedad") if r.get("diasAntiguedad") is not None else -1,
                reverse=True
            )
            sucursales_gestion.append({
                "idSucursal": grupo["idSucursal"],
                "nombre": grupo["nombre"],
                "bodegasIncluidas": grupo["bodegas"],
                "total": len(registros_suc),
                "registros": registros_suc,
            })
            total_gestion += len(registros_suc)

    wb.close()

    # ── Generar bodegas_gestion.json ──────────────────────────────────────────
    compartidas_registros.sort(
        key=lambda r: r.get("diasAntiguedad") if r.get("diasAntiguedad") is not None else -1,
        reverse=True
    )

    COMPARTIDAS_META = [
        {"id": 23, "simbolo": "CD",  "nombre": "Centro de Distribucion",        "idsucursal": "08"},
        {"id":  7, "simbolo": "XCD", "nombre": "CrossDock Centro Distribucion", "idsucursal": "08"},
        {"id": 27, "simbolo": "GCD", "nombre": "Gestion CD",                    "idsucursal": "08"},
        {"id": 73, "simbolo": "ICD", "nombre": "Ingreso Centro Distribucion",   "idsucursal": "08"},
        {"id": 26, "simbolo": "MCD", "nombre": "Mermas CD",                     "idsucursal": "08"},
        {"id": 54, "simbolo": "RCD", "nombre": "Recepcion Centro Distribucion", "idsucursal": "08"},
        {"id": 67, "simbolo": "TCD", "nombre": "Transito Centro Distribucion",  "idsucursal": "08"},
        {"id": 98, "simbolo": "BDP", "nombre": "Despacho Proveedor",            "idsucursal": "09"},
        {"id": 84, "simbolo": "REM", "nombre": "Remate",                        "idsucursal": "01"},
        {"id": 36, "simbolo": "MKT", "nombre": "Marketing",                     "idsucursal": "01"},
    ]

    data_gestion = {
        "generado": hoy.isoformat(),
        "fuente": "datos-bodegas.xlsm (ADODB SQL Server)",
        "sucursales": sucursales_gestion,
        "compartidas": {
            "bodegasIncluidas": COMPARTIDAS_META,
            "total": len(compartidas_registros),
            "registros": compartidas_registros,
        },
        "total": total_gestion,
    }

    if not anti_retroceso(OUT_GESTION, total_gestion, "bodegas_gestion"):
        sys.exit(1)

    OUT_GESTION.write_text(json.dumps(data_gestion, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OK] {OUT_GESTION.name}  total={total_gestion}")

    # ── Generar bodegas_ir_otras.json ─────────────────────────────────────────
    BODEGAS_IR_META = [g["bodegas"] for g in GRUPOS if g["json"] == "ir"][0]
    bodegas_ir.sort(
        key=lambda r: r.get("diasAntiguedad") if r.get("diasAntiguedad") is not None else -1,
        reverse=True
    )

    data_ir = {
        "generado": hoy.isoformat(),
        "fuente": "datos-bodegas.xlsm (ADODB SQL Server)",
        "bodegas": BODEGAS_IR_META,
        "total": total_ir,
        "registros": bodegas_ir,
    }

    if not anti_retroceso(OUT_IR, total_ir, "bodegas_ir_otras"):
        sys.exit(1)

    OUT_IR.write_text(json.dumps(data_ir, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] {OUT_IR.name}  total={total_ir}")
    print(f"\n[OK] Todo listo. Generado: {hoy.isoformat()}")


if __name__ == "__main__":
    main()
